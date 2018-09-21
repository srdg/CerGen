import os	
import shutil								
import openpyxl		
import datetime												
from PIL import Image													
from PIL import ImageFont
from PIL import ImageDraw										
#open Excel File									
ob=openpyxl.load_workbook('demodata.xlsx')								
#open Worksheet
sh1=ob['Form Responses 1']							
#set font to be used in text in Image
			
for i in range(2,sh1.max_row+1):
	#open Image file
	img=Image.open('Final.jpg')	
	W, H = img.size
	draw = ImageDraw.Draw(img)

	#getdata
	name = str(sh1.cell(row=i,column=2).value).upper()
	dept = str(sh1.cell(row=i,column=3).value).upper()
	begi = sh1.cell(row=i,column=4).value
	begi = begi.strftime('%m/%d/%Y').split('/')
	begi[0],begi[1]=begi[1],begi[0]
	begi = '/'.join(begi)
	endi = sh1.cell(row=i,column=5).value
	endi = endi.strftime('%m/%d/%Y').split('/')
	endi[0],endi[1] = endi[1],endi[0]
	endi = '/'.join(endi)
	task = str(sh1.cell(row=i,column=6).value).upper()	
	supervisor = str(sh1.cell(row=i,column=7).value).upper()	
	HOD = str(sh1.cell(row=i,column=9).value).upper()	

	#add text to Image
	font=ImageFont.truetype("Cambo-Regular.otf",size=60)	
	w, h = draw.textsize(name, font=font)
	draw.text(((W-w)/2,1400),name,(0,0,0),font=font)

	w, h = draw.textsize(dept, font=font)
	draw.text(((W-w)/2,1530),dept,(0,0,0),font=font)

	w, h = draw.textsize(begi, font=font)
	draw.text(((W-w)/2 - 150,1660),begi,(0,0,0),font=font)

	w, h = draw.textsize(endi, font=font)
	draw.text(((W-w)/2 + 250,1660),endi,(0,0,0),font=font)

	font=ImageFont.truetype("Cambo-Regular.otf",size=50)
	w, h = draw.textsize(task, font=font)
	draw.text(((W-w)/2,1808),task,(0,0,0),font=font)

	font=ImageFont.truetype("Cambo-Regular.otf",size=60)
	w, h = draw.textsize(supervisor, font=font)
	draw.text(((W-w)/5,2130),supervisor,(0,0,0),font=font)

	font=ImageFont.truetype("Cambo-Regular.otf",size=60)
	w, h = draw.textsize(HOD, font=font)
	draw.text(((W-w)/2,2130),HOD,(0,0,0),font=font)


	#save changes to Image file
	img.save(name+'.jpg')											
	img.close()
	#move Image file to new directory	
	try:
		if not os.path.exists("Certificates Generated Here"):
			os.makedirs("Certificates Generated Here")
		shutil.move(name+'.jpg', os.getcwd()+"/Certificates Generated Here")
	except:
		# traceback : file already exists
		os.rename(os.getcwd()+'/'+name+'.jpg', os.getcwd()+"/Certificates Generated Here"+'/'+name+'.jpg')
	print("Creating certificate for "+name+"....done.")
