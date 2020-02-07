import os	
import shutil								
import openpyxl		
import datetime
from PIL import Image													
from PIL import ImageFont
from PIL import ImageDraw										
#open Excel File									
ob=openpyxl.load_workbook('outdata.xlsx')								
#open Worksheet
sh1=ob['Form Responses 1']
#set font to be used in text in Image

for i in range(2,sh1.max_row+1):
	#open Image file
	img=Image.open('final.jpg')	
	W, H = img.size
	draw = ImageDraw.Draw(img)

	#getdata
	jgec_dept = 'DEPARTMENT OF '+str(sh1.cell(row=i,column=3).value).upper()
	name_intern = str(sh1.cell(row=i,column=4).value).upper()
	clg_dept = str(sh1.cell(row=i,column=5).value).upper()
	college = str(sh1.cell(row=i,column=6).value).upper()

	start_date = sh1.cell(row=i,column=7).value
	end_date = sh1.cell(row=i,column=8).value
	start_date = start_date.strftime("%m/%d/%Y").split('/')
	start_date[0],start_date[1] = start_date[1],start_date[0]
	end_date = end_date.strftime("%m/%d/%Y").split('/')
	end_date[0],end_date[1] = end_date[1],end_date[0]
	start_date = '/'.join(start_date)
	end_date = '/'.join(end_date)
	
	title = str(sh1.cell(row=i,column=9).value).upper()
	supervisor = str(sh1.cell(row=i,column=10).value).upper()
	head_of_dept = str(sh1.cell(row=i,column=11).value).upper()


	font=ImageFont.truetype("Bentham.otf",size=80)	
	w, h = draw.textsize(jgec_dept, font=font)
	draw.text((((W-w)/2)+50,970),jgec_dept,(0,0,0),font=font)


	font=ImageFont.truetype("Cambo-Regular.otf",size=60)	
	w, h = draw.textsize(name_intern, font=font)
	draw.text((((W-w)/2)+50,1190),name_intern,(0,0,0),font=font)

	font=ImageFont.truetype("Cambo-Regular.otf",size=50)	
	w, h = draw.textsize('from', font=font)
	draw.text((((W-w)/2)+50,1250),'from',(0,0,0),font=font)

	font=ImageFont.truetype("Cambo-Regular.otf",size=60)	
	w, h = draw.textsize(clg_dept, font=font)
	draw.text((((W-w)/2)+50,1300),clg_dept,(0,0,0),font=font)

	font=ImageFont.truetype("Cambo-Regular.otf",size=50)	
	w, h = draw.textsize('of', font=font)
	draw.text((((W-w)/2)+50,1360),'of',(0,0,0),font=font)

	font=ImageFont.truetype("Cambo-Regular.otf",size=60)	
	w, h = draw.textsize(college, font=font)
	draw.text((((W-w)/2)+50,1420),college,(0,0,0),font=font)

	w, h = draw.textsize(start_date, font=font)
	draw.text(((W-w)/2 - 150,1570),start_date,(0,0,0),font=font)

	w, h = draw.textsize(end_date, font=font)
	draw.text(((W-w)/2 + 250,1570),end_date,(0,0,0),font=font)

	font=ImageFont.truetype("Cambo-Regular.otf",size=50)
	w, h = draw.textsize(title, font=font)
	draw.text(((W-w)/2 + 50,1710),title,(0,0,0),font=font)


	w, h = draw.textsize(supervisor, font=font)
	draw.text(((W-w)/8,2170),supervisor,(0,0,0),font=font)

	w, h = draw.textsize(head_of_dept, font=font)
	draw.text(((W-w)/3 + 130,2170),head_of_dept,(0,0,0),font=font)

	img.save(name_intern+'.jpg')											
	img.close()
	img.close()
	#move Image file to new directory	
	try:
		if not os.path.exists("Outdoor certificates Generated Here"):
			os.makedirs("Outdoor certificates Generated Here")
		shutil.move(name_intern+'.jpg', os.getcwd()+"/Outdoor certificates Generated Here")
	except:
		# traceback : file already exists
		os.rename(os.getcwd()+'/'+name_intern+'.jpg', os.getcwd()+"/Outdoor certificates Generated Here"+'/'+name_intern+'.jpg')
	print("Creating certificate for " + name_intern + "....done.")
